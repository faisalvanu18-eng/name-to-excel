from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)

EXCEL_FILE = "records.xlsx"
SHEET_NAME = "Records"


def init_excel_if_needed():
    """Create Excel file with headers if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["Name", "Date", "Time"])  # headers
        wb.save(EXCEL_FILE)


def save_record(name: str):
    init_excel_if_needed()

    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    ws.append([name, date_str, time_str])
    wb.save(EXCEL_FILE)


@app.route("/", methods=["GET"])
def home():
    return render_template("index.html")


@app.route("/submit", methods=["POST"])
def submit():
    name = request.form.get("name", "").strip()
    if not name:
        return redirect(url_for("home"))

    save_record(name)
    return redirect(url_for("success"))


@app.route("/success", methods=["GET"])
def success():
    return render_template("success.html")


# ✅ View all saved records in browser
@app.route("/records", methods=["GET"])
def records():
    init_excel_if_needed()

    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header row
        rows.append(row)

    return render_template("records.html", rows=rows)


if __name__ == "__main__":
    # Local run (PC): python server.py
    # On Render (deployed), gunicorn runs the app
    app.run(debug=True)
